"""
Network Connections Dashboard (Pipeline + Railway + Truck)

Three tabs over the same node_metrics_150.xlsx:

1. "Pipeline size classes" - the three CO2 pipeline size classes
   (CO2_Pipeline_small/medium/large - see main_italy.py's
   pipeline_size_class_max_capacity_t_h and
   data_process/updated_network/pipeline_capex_per_arc_calculator.py::
   SIZE_CLASS_MASSFLOW_RANGES_KG_S) all reuse the exact same physical pipeline
   connectivity by default: a 'large' pipeline (calibrated for ~133-470 kg/s)
   is currently just as buildable on an arc that only ever carries a small
   emitter's flow as a 'small' one is. This tab lets you remove specific arcs
   from an individual size class instead of every class sharing the same flat
   connectivity.

   Workflow:
     a. Pick a size class in the sidebar (or just click cells in the table -
        see below).
     b. Click a "Small"/"Medium"/"Large" cell in the arc table to toggle that
        arc on/off for that specific class - every cell saves immediately.
        Alternatively, pick a class with the radio buttons and click an arc
        directly on the map to toggle it for the selected class.
     c. main_italy.py picks up your edits automatically on the next run (via
        defined_functions.load_pipeline_class_connection_matrix reading
        pipeline_size_class_connections.xlsx) - no code changes needed.

   An arc not touched here stays enabled for every class (identical to the
   flat, pre-size-class behaviour), so this tab only ever narrows
   connectivity, never adds arcs beyond what node_metrics_150.xlsx's
   'pipeline' sheet already allows.

2. "Railway network" - read-only viewer over the 'railway' sheet's arcs.
   node_metrics_150.xlsx is hand-maintained and this tab never writes to it.
   It exists to make the rail network's actual shape visible at a glance:
     - which arcs exist, and whether each is one-way or bidirectional in the
       raw data (update_network_connection_matrix() in defined_functions.py
       converts any value > 0 straight into a directed solver constraint with
       no symmetrization, so a one-way gap here is a real modelling gap, not
       just a data-entry quirk - see
       italy_data/geographical_feature/train_stations_analysis/README.md #1);
     - which 'Transport' nodes (stations) are actually reachable by rail at
       all, vs. defined-but-orphaned (0 arcs);
     - an optional overlay of the candidate replacement/new stations from the
       station audit in train_stations_analysis/README.md, so you can compare
       today's (often wrong or defunct) station placements against real
       freight-equipped facilities before editing the sheet by hand.

3. "Truck network" - the only tab that writes anything back to
   node_metrics_150.xlsx. Lets you add a new truck arc (a node pair not yet
   connected in the 'truck' sheet), get its distance from live OSM routing
   (reusing data_process/updated_network/truck_routing.py's own
   download_od_subgraph/route_distance_km - restricted to truck-suitable
   road classes) or type one in by hand, then save it directly into the
   'truck' sheet's matching cell.

   This writes straight into the sheet (not a side override file) because
   defined_functions.compute_opex_var_arcs and .update_capex_gamma2_per_arc
   independently re-read node_metrics_150.xlsx's 'truck' sheet by path when
   main_italy.py runs - an override layer only this dashboard knew about
   would desync from the cost model. Truck cost is a pure function of
   distance (no manual per-arc curation step like the pipeline mass-flow
   editor's gammas), so this one write is also sufficient: the next
   main_italy.py run re-reads the file fresh and picks up the new arc's
   topology, opex and capex automatically, no code changes needed.

   Only the 'truck' sheet's specific cell is touched (via openpyxl, not a
   full-sheet pandas rewrite) - every other sheet/cell is left as-is. The
   whole file is backed up once per dashboard session (node_metrics_150.xlsx.bak-
   <timestamp>) before the first write. A separate audit log
   (truck_arcs_added_via_dashboard.csv, not read by main_italy.py) records
   every arc added here, and is what distinguishes "added via dashboard"
   arcs from the original 21 ArcGIS-computed ones on the map/table.

Run with: python network_connections_dashboard.py, then open
http://127.0.0.1:8052
"""

import datetime
import math
import shutil
import sys
from pathlib import Path

import openpyxl
import pandas as pd
import dash
from dash import dcc, html, Input, Output, State, dash_table
import dash_bootstrap_components as dbc
import plotly.graph_objects as go

# ==========================================
# 1. PATH SETUP
# ==========================================
# This script lives inside italy_data/geographical_feature itself, alongside
# node_metrics_150.xlsx, so paths are relative to that.
SCRIPT_DIR = Path(__file__).resolve().parent
DATA_PATH = SCRIPT_DIR.parent
SOUTHERN_EUROPE_DIR = DATA_PATH.parent
CAPEX_METRICS_DIR = DATA_PATH / "network_capex_metrics"

# Must match main_italy.py's node_metrics_suffix (default 150) - this is the
# exact 'pipeline'/'railway'/'truck' connectivity/distance matrices
# main_italy.py reads before applying the per-class overrides the pipeline
# tab curates (or, for truck, straight as-is).
NODE_METRICS_PATH = SCRIPT_DIR / "node_metrics_150.xlsx"
OVERRIDES_PATH = CAPEX_METRICS_DIR / "pipeline_size_class_connections.xlsx"
TRUCK_AUDIT_LOG_PATH = SCRIPT_DIR / "truck_arcs_added_via_dashboard.csv"

if not NODE_METRICS_PATH.exists():
    raise FileNotFoundError(f"Critical Error: Could not locate {NODE_METRICS_PATH}")

# Reuse the actual OSM-routing implementation rather than duplicating it.
sys.path.append(str(SOUTHERN_EUROPE_DIR / "data_process" / "updated_network"))
from truck_routing import download_od_subgraph, route_distance_km  # noqa: E402

SIZE_CLASSES = ["small", "medium", "large"]

# Reference only, kept in sync with pipeline_capex_per_arc_calculator.py::
# SIZE_CLASS_MASSFLOW_RANGES_KG_S and main_italy.py::
# pipeline_size_class_max_capacity_t_h - shown in the sidebar/table to help
# judge which arcs make sense for which class.
SIZE_CLASS_RANGES_KG_S = {"small": (3.1, 29.0), "medium": (29.0, 133.0), "large": (133.0, 470.0)}
SIZE_CLASS_MAX_CAPACITY_T_H = {"small": 104.4, "medium": 478.8, "large": 1692.0}

SECONDS_PER_YEAR = 365.25 * 24 * 3600
TONNES_TO_KG = 1000  # 'annual_flux' is in tonnes CO2/year despite the column name

# Candidate replacement/new rail stations from the station audit - see
# train_stations_analysis/README.md for full sourcing and reasoning. Purely a
# visual overlay; none of this is written back to node_metrics_150.xlsx.
CANDIDATE_STATIONS = [
    {"replaces": 3, "replaces_name": "Trino Vercellese", "name": "Torino Orbassano",
     "lon": 7.5712, "lat": 45.0311,
     "note": "Active RFI + FS Logistica freight/intermodal yard. Trino Vercellese is passenger-only today."},
    {"replaces": 15, "replaces_name": "S.Giuliano Milanese", "name": "Segrate",
     "lon": 9.299, "lat": 45.481,
     "note": "Active RFI freight terminal (Terminali Italia), almost same location as today's node."},
    {"replaces": 15, "replaces_name": "S.Giuliano Milanese", "name": "Melzo",
     "lon": 9.410, "lat": 45.496,
     "note": "Alternative to Segrate: larger Contship/Sogemar intermodal terminal, 13 km further east."},
    {"replaces": 38, "replaces_name": "Venezia", "name": "Venezia Marghera Scalo",
     "lon": 12.2418, "lat": 45.4671,
     "note": "Confirmed freight-only station; today's node sits imprecisely on the causeway."},
    {"replaces": 42, "replaces_name": "Eni S.p.A Casalborsetti", "name": "Ravenna / TCRavenna",
     "lon": 12.25, "lat": 44.42,
     "note": "Address-level only, confirm on a map. Node 42 looks mislabeled as a station (it's the CCUS source)."},
    {"replaces": 44, "replaces_name": "Modena-H", "name": "Marzaglia Intermodal Terminal",
     "lon": 10.8412, "lat": 44.6534,
     "note": "Modena's freight yard was relocated here (2018-2021, EUR110M, Terminali Italia). Today's node has wrong coordinates."},
    {"replaces": None, "replaces_name": None, "name": "Trieste Campo Marzio Smistamento",
     "lon": 13.7545, "lat": 45.6406,
     "note": "New station: active freight marshalling yard, ~4 km from the Trieste incinerator vs. today's 162 km detour via Venezia."},
]

# ==========================================
# 2. LOAD NETWORK DATA ONCE AT STARTUP
# ==========================================
print("=" * 80)
print("Loading network data for the network connections dashboard...")
print("=" * 80)

_nodes_raw = pd.read_excel(NODE_METRICS_PATH, sheet_name="nodes", index_col=0)
BASE_PIPELINE = pd.read_excel(NODE_METRICS_PATH, sheet_name="pipeline", index_col=0)
BASE_PIPELINE.columns = BASE_PIPELINE.columns.astype(int)
BASE_RAILWAY = pd.read_excel(NODE_METRICS_PATH, sheet_name="railway", index_col=0)
BASE_RAILWAY.columns = BASE_RAILWAY.columns.astype(int)

# A node_id can be shared by several co-located facility rows (e.g. one node
# has both a Waste and a Cement row) - collapse those for display and sum
# their emissions, same convention as massflow_editor_dashboard.py.
_nodes_raw["node_type"] = _nodes_raw["node_type"].astype(str).str.strip()
NODES_RAW = _nodes_raw.groupby(_nodes_raw.index).agg(
    node_name=("node_name", "first"),
    longitude=("longitude", "first"),
    latitude=("latitude", "first"),
    node_type=("node_type", lambda s: "+".join(sorted(set(s)))),
    annual_flux=("annual_flux", "sum"),
)
NODES_RAW.index.name = "node_id"
NODES_RAW["emission_kg_s"] = NODES_RAW["annual_flux"].fillna(0.0) * TONNES_TO_KG / SECONDS_PER_YEAR
NODE_NAME = NODES_RAW["node_name"].to_dict()

ALL_NODE_IDS = sorted(set(BASE_PIPELINE.index) | set(BASE_PIPELINE.columns))

# All directed arcs that physically exist (value > 0). The underlying matrix
# is NOT symmetric - e.g. 1->3 can be connected while 3->1 is not - so each
# direction is tracked and toggled independently.
POSSIBLE_ARCS = [
    (f, t) for f in BASE_PIPELINE.index for t in BASE_PIPELINE.columns
    if pd.notna(BASE_PIPELINE.loc[f, t]) and BASE_PIPELINE.loc[f, t] > 0
]

# Same idea for the railway sheet.
POSSIBLE_RAIL_ARCS = [
    (f, t) for f in BASE_RAILWAY.index for t in BASE_RAILWAY.columns
    if pd.notna(BASE_RAILWAY.loc[f, t]) and BASE_RAILWAY.loc[f, t] > 0
]
RAIL_ARC_SET = set(POSSIBLE_RAIL_ARCS)
RAIL_CONNECTED_NODES = {n for pair in POSSIBLE_RAIL_ARCS for n in pair}
ORPHANED_TRANSPORT_NODES = [
    nid for nid in NODES_RAW.index
    if "Transport" in NODES_RAW.loc[nid, "node_type"] and nid not in RAIL_CONNECTED_NODES
]

print(f"Loaded {len(NODES_RAW)} nodes, {len(POSSIBLE_ARCS)} directed pipeline arcs, "
      f"{len(POSSIBLE_RAIL_ARCS)} directed railway arcs.")
print(f"Orphaned 'Transport' nodes (0 railway arcs): {ORPHANED_TRANSPORT_NODES}")
print("=" * 80)


# ==========================================
# 3. PIPELINE OVERRIDE FILE HELPERS
# ==========================================
def load_class_mask(size_class):
    """1 = enabled, 0 = disabled, for every (from,to) node_id pair. Defaults
    to all-enabled if the override file/sheet doesn't exist yet - matches
    main_italy.py / load_pipeline_class_connection_matrix's default."""
    default = pd.DataFrame(1, index=ALL_NODE_IDS, columns=ALL_NODE_IDS)
    if not OVERRIDES_PATH.exists():
        return default
    try:
        xl = pd.ExcelFile(OVERRIDES_PATH)
        if size_class not in xl.sheet_names:
            return default
        df = pd.read_excel(OVERRIDES_PATH, sheet_name=size_class, index_col=0)
        df.index = df.index.astype(int)
        df.columns = df.columns.astype(int)
    except Exception as e:
        print(f"Warning: could not load '{size_class}' override sheet: {e}")
        return default
    return df.reindex(index=ALL_NODE_IDS, columns=ALL_NODE_IDS, fill_value=1).fillna(1).astype(int)


def save_class_mask(size_class, mask_df):
    """Upsert one class's mask sheet into pipeline_size_class_connections.xlsx,
    leaving every other class's sheet untouched."""
    OVERRIDES_PATH.parent.mkdir(parents=True, exist_ok=True)
    all_sheets = {}
    if OVERRIDES_PATH.exists():
        all_sheets = pd.read_excel(OVERRIDES_PATH, sheet_name=None, index_col=0)
    all_sheets[size_class] = mask_df
    with pd.ExcelWriter(OVERRIDES_PATH, engine="openpyxl") as writer:
        for name, df in all_sheets.items():
            df.to_excel(writer, sheet_name=name, index=True)


def toggle_arc(size_class, from_node, to_node):
    mask_df = load_class_mask(size_class)
    current = int(mask_df.loc[from_node, to_node])
    new_val = 0 if current == 1 else 1
    mask_df.loc[from_node, to_node] = new_val
    save_class_mask(size_class, mask_df)
    return new_val


def set_all_arcs(size_class, value):
    mask_df = load_class_mask(size_class)
    for f, t in POSSIBLE_ARCS:
        mask_df.loc[f, t] = value
    save_class_mask(size_class, mask_df)


def build_arcs_table_data():
    masks = {sc: load_class_mask(sc) for sc in SIZE_CLASSES}
    rows = []
    for f, t in POSSIBLE_ARCS:
        row = {
            "from_name": NODE_NAME.get(f, f"Node {f}"),
            "from_node": f,
            "to_name": NODE_NAME.get(t, f"Node {t}"),
            "to_node": t,
            "distance_km": round(float(BASE_PIPELINE.loc[f, t]), 2),
            "from_flow_kg_s": round(float(NODES_RAW.loc[f, "emission_kg_s"]), 3) if f in NODES_RAW.index else None,
        }
        for sc in SIZE_CLASSES:
            row[sc] = "✅" if bool(masks[sc].loc[f, t]) else "❌"
        rows.append(row)
    return rows


# ==========================================
# 4. RAILWAY TABLE HELPERS (read-only)
# ==========================================
def build_rail_arcs_table_data():
    rows = []
    for f, t in POSSIBLE_RAIL_ARCS:
        rows.append({
            "from_name": NODE_NAME.get(f, f"Node {f}"),
            "from_node": f,
            "to_name": NODE_NAME.get(t, f"Node {t}"),
            "to_node": t,
            "distance_km": round(float(BASE_RAILWAY.loc[f, t]), 2),
            "direction": "Bidirectional" if (t, f) in RAIL_ARC_SET else "One-way only",
            "from_type": NODES_RAW.loc[f, "node_type"] if f in NODES_RAW.index else "?",
            "to_type": NODES_RAW.loc[t, "node_type"] if t in NODES_RAW.index else "?",
        })
    return rows


def build_orphaned_stations_list():
    return [
        html.Li(f"#{nid} - {NODE_NAME.get(nid, 'Unknown')}", className="text-danger")
        for nid in ORPHANED_TRANSPORT_NODES
    ]


# ==========================================
# 4B. TRUCK TAB HELPERS - routing, audit log, xlsx write
# ==========================================
NODE_DROPDOWN_OPTIONS = [
    {"label": f"{row.node_name} (#{nid})", "value": int(nid)}
    for nid, row in NODES_RAW.sort_values("node_name").iterrows()
]

TRUCK_AUDIT_COLUMNS = ["timestamp", "from_node", "from_name", "to_node", "to_name",
                       "distance_km", "method", "previous_value"]

# Set once per dashboard process, the first time write_truck_arc() runs.
_truck_backup_done = False


def _ensure_truck_backup():
    """Copy node_metrics_150.xlsx to a timestamped .bak file once per
    dashboard session, before the very first write to it - so a mistaken
    save (bad node pair, bad distance) is trivially reversible."""
    global _truck_backup_done
    if _truck_backup_done:
        return
    ts = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
    backup_path = NODE_METRICS_PATH.with_name(f"{NODE_METRICS_PATH.name}.bak-{ts}")
    shutil.copy2(NODE_METRICS_PATH, backup_path)
    _truck_backup_done = True
    print(f"Backed up {NODE_METRICS_PATH.name} -> {backup_path.name}")


def load_truck_matrix() -> pd.DataFrame:
    """Always re-read from disk (unlike BASE_PIPELINE/BASE_RAILWAY, which are
    loaded once at startup) since this is the one sheet the dashboard itself
    writes to - callers need the latest values right after a save."""
    df = pd.read_excel(NODE_METRICS_PATH, sheet_name="truck", index_col=0)
    df.columns = df.columns.astype(int)
    return df


def get_truck_arcs(df: pd.DataFrame | None = None) -> list[tuple[int, int]]:
    if df is None:
        df = load_truck_matrix()
    return [(f, t) for f in df.index for t in df.columns if pd.notna(df.loc[f, t]) and df.loc[f, t] > 0]


def load_truck_audit_log() -> pd.DataFrame:
    if TRUCK_AUDIT_LOG_PATH.exists():
        return pd.read_csv(TRUCK_AUDIT_LOG_PATH)
    return pd.DataFrame(columns=TRUCK_AUDIT_COLUMNS)


def append_truck_audit_log(from_node, to_node, distance_km, method, previous_value):
    row = {
        "timestamp": datetime.datetime.now().isoformat(timespec="seconds"),
        "from_node": from_node, "from_name": NODE_NAME.get(from_node, from_node),
        "to_node": to_node, "to_name": NODE_NAME.get(to_node, to_node),
        "distance_km": distance_km, "method": method,
        "previous_value": previous_value,
    }
    pd.DataFrame([row], columns=TRUCK_AUDIT_COLUMNS).to_csv(
        TRUCK_AUDIT_LOG_PATH, mode="a", header=not TRUCK_AUDIT_LOG_PATH.exists(), index=False
    )


def dashboard_added_truck_arcs() -> dict:
    """(from_node, to_node) -> most recent audit-log row, for every arc ever
    saved via this dashboard (used to color/label them differently from the
    original 21 ArcGIS arcs - it does NOT mean the arc is still non-zero if
    someone later zeroed it out by hand elsewhere)."""
    log = load_truck_audit_log()
    if log.empty:
        return {}
    out = {}
    for row in log.itertuples():
        out[(int(row.from_node), int(row.to_node))] = row
    return out


def _truck_sheet_cell_maps(ws):
    """Map node_id -> row/column index for the 'truck' worksheet, reading
    the header row and index column directly via openpyxl (not pandas) so
    write_truck_arc() can edit a single cell without touching anything else
    in the workbook."""
    row_map, col_map = {}, {}
    for c in range(2, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is not None:
            col_map[int(v)] = c
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is not None:
            row_map[int(v)] = r
    return row_map, col_map


def get_current_truck_value(from_node: int, to_node: int):
    wb = openpyxl.load_workbook(NODE_METRICS_PATH, data_only=True, read_only=True)
    try:
        ws = wb["truck"]
        row_map, col_map = _truck_sheet_cell_maps(ws)
        return ws.cell(row=row_map[from_node], column=col_map[to_node]).value
    finally:
        wb.close()


def write_truck_arc(from_node: int, to_node: int, distance_km: float, method: str) -> float | None:
    """Writes one directed truck-arc distance directly into
    node_metrics_150.xlsx's 'truck' sheet, at the (from_node, to_node) cell
    only - every other sheet and cell is left untouched. Backs up the whole
    file once per session first (see _ensure_truck_backup). Returns the
    previous value (None/0 if the arc didn't exist before) and appends an
    entry to the dashboard's own audit log.

    Note: openpyxl round-trips the workbook (load -> edit one cell -> save);
    this preserves values/formatting on every normal sheet in this file, but
    is a reason on top of the backup to spot-check the file after a save if
    it ever grows unusual features (e.g. embedded charts) the other sheets
    don't currently have.
    """
    _ensure_truck_backup()
    wb = openpyxl.load_workbook(NODE_METRICS_PATH)
    try:
        ws = wb["truck"]
        row_map, col_map = _truck_sheet_cell_maps(ws)
        if from_node not in row_map or to_node not in col_map:
            raise ValueError(f"Node {from_node} or {to_node} not found in the 'truck' sheet's headers.")
        cell = ws.cell(row=row_map[from_node], column=col_map[to_node])
        previous_value = cell.value
        cell.value = float(distance_km)
        wb.save(NODE_METRICS_PATH)
    finally:
        wb.close()
    append_truck_audit_log(from_node, to_node, distance_km, method, previous_value)
    return previous_value


def compute_truck_distance(from_node: int, to_node: int):
    """Live OSM-routed distance (km) for one directed pair, via
    truck_routing.py's own download+shortest-path helpers (same
    truck-suitable road-class filter it uses for the reference validation
    set). Returns (distance_km_or_None, status_message)."""
    if from_node is None or to_node is None:
        return None, "Select both a From and To node first."
    if from_node == to_node:
        return None, "From and To must be different nodes."
    fr, to = NODES_RAW.loc[from_node], NODES_RAW.loc[to_node]
    G = download_od_subgraph(fr.longitude, fr.latitude, to.longitude, to.latitude)
    if G is None:
        return None, ("Could not download the OSM road network for this pair (connection issue). "
                       "Try again, or enter a distance manually below.")
    d = route_distance_km(G, fr.longitude, fr.latitude, to.longitude, to.latitude)
    if d is None:
        return None, ("No truck-suitable road path found between these two nodes on OSM. "
                       "Enter a distance manually below if you have one from another source.")
    return d, f"Computed via OSM routing: {d:.2f} km"


def build_truck_arcs_table_data():
    df = load_truck_matrix()
    arcs = get_truck_arcs(df)
    added = dashboard_added_truck_arcs()
    rows = []
    for f, t in arcs:
        entry = added.get((f, t))
        rows.append({
            "from_name": NODE_NAME.get(f, f"Node {f}"), "from_node": f,
            "to_name": NODE_NAME.get(t, f"Node {t}"), "to_node": t,
            "distance_km": round(float(df.loc[f, t]), 2),
            "source": f"Dashboard ({entry.method})" if entry is not None else "Original (ArcGIS)",
            "added_at": entry.timestamp if entry is not None else "",
        })
    return rows


# ==========================================
# 5. MAP RENDERING - SHARED HELPERS
# ==========================================
def _bearing(lat1, lon1, lat2, lon2):
    lat1, lon1, lat2, lon2 = map(math.radians, [lat1, lon1, lat2, lon2])
    dlon = lon2 - lon1
    x = math.sin(dlon) * math.cos(lat2)
    y = math.cos(lat1) * math.sin(lat2) - math.sin(lat1) * math.cos(lat2) * math.cos(dlon)
    return (math.degrees(math.atan2(x, y)) + 360) % 360


def _arrow_tip(lat1, lon1, lat2, lon2, offset_frac=0.2):
    tip_lat = lat2 + offset_frac * (lat1 - lat2)
    tip_lon = lon2 + offset_frac * (lon1 - lon2)
    return tip_lat, tip_lon


def _base_geo_layout(fig):
    fig.update_layout(
        geo=dict(scope="europe", center=dict(lon=12.6, lat=44.8), projection_scale=6.5,
                 showland=True, landcolor="#f9f9f9", countrycolor="#bdc3c7"),
        margin=dict(l=0, r=0, t=0, b=0), showlegend=False,
        hovermode="closest",
        clickmode="event",
    )
    return fig


ENABLED_COLOR = "rgba(39, 174, 96, 0.85)"
DISABLED_COLOR = "rgba(192, 57, 43, 0.55)"
ENABLED_SOLID = "#27ae60"
DISABLED_SOLID = "#c0392b"


# ==========================================
# 6. MAP RENDERING - PIPELINE TAB
# ==========================================
def generate_pipeline_map_figure(size_class):
    mask_df = load_class_mask(size_class)
    fig = go.Figure()

    arrow_lons, arrow_lats, arrow_angles, arrow_hovers, arrow_colors, arrow_customdata = [], [], [], [], [], []
    hit_lons, hit_lats, hit_hovers, hit_customdata = [], [], [], []

    for f, t in POSSIBLE_ARCS:
        if f not in NODES_RAW.index or t not in NODES_RAW.index:
            continue
        node_a, node_b = NODES_RAW.loc[f], NODES_RAW.loc[t]
        enabled = bool(mask_df.loc[f, t])
        color = ENABLED_COLOR if enabled else DISABLED_COLOR
        solid = ENABLED_SOLID if enabled else DISABLED_SOLID

        hover_txt = (
            f"<b>{node_a.node_name}</b> (#{f}) &rarr; <b>{node_b.node_name}</b> (#{t})<br>"
            f"Distance: {BASE_PIPELINE.loc[f, t]:.2f} km<br>"
            f"'{size_class}' status: {'ENABLED' if enabled else 'DISABLED'} (click to toggle)"
        )

        fig.add_trace(go.Scattergeo(
            lon=[node_a.longitude, node_b.longitude],
            lat=[node_a.latitude, node_b.latitude],
            mode="lines",
            line=dict(width=2.5, color=color),
            hoverinfo="skip",
        ))

        hit_lons += [node_a.longitude, node_b.longitude, None]
        hit_lats += [node_a.latitude, node_b.latitude, None]
        hit_hovers += [hover_txt, hover_txt, hover_txt]
        hit_customdata += [[int(f), int(t)], [int(f), int(t)], [int(f), int(t)]]

        tip_lat, tip_lon = _arrow_tip(node_a.latitude, node_a.longitude, node_b.latitude, node_b.longitude)
        bearing = _bearing(node_a.latitude, node_a.longitude, node_b.latitude, node_b.longitude)
        arrow_lons.append(tip_lon)
        arrow_lats.append(tip_lat)
        arrow_angles.append(bearing)
        arrow_hovers.append(hover_txt)
        arrow_colors.append(solid)
        arrow_customdata.append([int(f), int(t)])

    if hit_lons:
        fig.add_trace(go.Scattergeo(
            lon=hit_lons, lat=hit_lats, mode="lines",
            line=dict(width=14, color="rgba(0,0,0,0.001)"),
            hoverinfo="text", hovertext=hit_hovers, customdata=hit_customdata,
            name="Arc (click target)",
        ))

    if arrow_lons:
        fig.add_trace(go.Scattergeo(
            lon=arrow_lons, lat=arrow_lats, mode="markers",
            marker=dict(symbol="arrow", size=11, color=arrow_colors, angle=arrow_angles, line=dict(width=0)),
            hoverinfo="text", hovertext=arrow_hovers, customdata=arrow_customdata, name="Direction",
        ))

    node_hover = [
        f"{row.node_name} (#{nid})<br>Type: {row.node_type}<br>Flow: {row.emission_kg_s:.3f} kg/s"
        for nid, row in NODES_RAW.iterrows()
    ]
    fig.add_trace(go.Scattergeo(
        lon=NODES_RAW["longitude"], lat=NODES_RAW["latitude"],
        mode="markers+text",
        text=NODES_RAW.index.astype(str),
        textposition="top right",
        marker=dict(size=10, color="#2c3e50", line=dict(width=1.2, color="#ffffff")),
        hoverinfo="text", hovertext=node_hover, name="Nodes",
    ))

    fig.update_layout(uirevision="pipeline-class-map")
    return _base_geo_layout(fig)


# ==========================================
# 7. MAP RENDERING - RAILWAY TAB
# ==========================================
BIDIRECTIONAL_COLOR = "#16a085"
ONEWAY_COLOR = "#e67e22"
ORPHAN_RING_COLOR = "#c0392b"
CANDIDATE_COLOR = "#8e44ad"

NODE_TYPE_COLORS = {
    "Waste": "#7f8c8d",
    "Cement": "#8e5a2f",
    "Refining": "#2980b9",
    "Transport": "#f39c12",
    "Storage": "#9b59b6",
    "Other": "#16a085",
}


def _node_color(node_type):
    primary = str(node_type).split("+")[0]
    return NODE_TYPE_COLORS.get(primary, "#2c3e50")


def generate_rail_map_figure(show_candidates):
    fig = go.Figure()

    arrow_lons, arrow_lats, arrow_angles, arrow_hovers, arrow_colors, arrow_customdata = [], [], [], [], [], []
    hit_lons, hit_lats, hit_hovers, hit_customdata = [], [], [], []

    for f, t in POSSIBLE_RAIL_ARCS:
        if f not in NODES_RAW.index or t not in NODES_RAW.index:
            continue
        node_a, node_b = NODES_RAW.loc[f], NODES_RAW.loc[t]
        bidirectional = (t, f) in RAIL_ARC_SET
        color = BIDIRECTIONAL_COLOR if bidirectional else ONEWAY_COLOR

        hover_txt = (
            f"<b>{node_a.node_name}</b> (#{f}) &rarr; <b>{node_b.node_name}</b> (#{t})<br>"
            f"Distance: {BASE_RAILWAY.loc[f, t]:.2f} km<br>"
            f"{'Bidirectional' if bidirectional else 'One-way only (reverse arc missing in the sheet)'}"
        )

        fig.add_trace(go.Scattergeo(
            lon=[node_a.longitude, node_b.longitude],
            lat=[node_a.latitude, node_b.latitude],
            mode="lines",
            line=dict(width=3, color=color),
            hoverinfo="skip",
        ))

        hit_lons += [node_a.longitude, node_b.longitude, None]
        hit_lats += [node_a.latitude, node_b.latitude, None]
        hit_hovers += [hover_txt, hover_txt, hover_txt]
        hit_customdata += [[int(f), int(t)], [int(f), int(t)], [int(f), int(t)]]

        tip_lat, tip_lon = _arrow_tip(node_a.latitude, node_a.longitude, node_b.latitude, node_b.longitude)
        bearing = _bearing(node_a.latitude, node_a.longitude, node_b.latitude, node_b.longitude)
        arrow_lons.append(tip_lon)
        arrow_lats.append(tip_lat)
        arrow_angles.append(bearing)
        arrow_hovers.append(hover_txt)
        arrow_colors.append(color)
        arrow_customdata.append([int(f), int(t)])

    if hit_lons:
        fig.add_trace(go.Scattergeo(
            lon=hit_lons, lat=hit_lats, mode="lines",
            line=dict(width=14, color="rgba(0,0,0,0.001)"),
            hoverinfo="text", hovertext=hit_hovers, customdata=hit_customdata,
            name="Arc",
        ))

    if arrow_lons:
        fig.add_trace(go.Scattergeo(
            lon=arrow_lons, lat=arrow_lats, mode="markers",
            marker=dict(symbol="arrow", size=12, color=arrow_colors, angle=arrow_angles, line=dict(width=0)),
            hoverinfo="text", hovertext=arrow_hovers, customdata=arrow_customdata, name="Direction",
        ))

    # All nodes, colored by type.
    node_colors = [_node_color(t) for t in NODES_RAW["node_type"]]
    node_hover = [
        f"{row.node_name} (#{nid})<br>Type: {row.node_type}<br>Flow: {row.emission_kg_s:.3f} kg/s"
        + ("<br><b>ORPHANED - no railway arcs</b>" if nid in ORPHANED_TRANSPORT_NODES else "")
        for nid, row in NODES_RAW.iterrows()
    ]
    node_sizes = [
        14 if "Transport" in t else 9
        for t in NODES_RAW["node_type"]
    ]
    fig.add_trace(go.Scattergeo(
        lon=NODES_RAW["longitude"], lat=NODES_RAW["latitude"],
        mode="markers+text",
        text=NODES_RAW.index.astype(str),
        textposition="top right",
        marker=dict(size=node_sizes, color=node_colors, line=dict(width=1.2, color="#ffffff")),
        hoverinfo="text", hovertext=node_hover, name="Nodes",
    ))

    # Highlight orphaned Transport nodes with a red ring on top.
    if ORPHANED_TRANSPORT_NODES:
        orphan_rows = NODES_RAW.loc[ORPHANED_TRANSPORT_NODES]
        fig.add_trace(go.Scattergeo(
            lon=orphan_rows["longitude"], lat=orphan_rows["latitude"],
            mode="markers",
            marker=dict(size=22, color="rgba(0,0,0,0)", line=dict(width=3, color=ORPHAN_RING_COLOR)),
            hoverinfo="skip", name="Orphaned station",
        ))

    if show_candidates and CANDIDATE_STATIONS:
        cand_hover = [
            f"<b>Candidate: {c['name']}</b><br>"
            + (f"Replaces #{c['replaces']} ({c['replaces_name']})<br>" if c["replaces"] else "New station<br>")
            + c["note"]
            for c in CANDIDATE_STATIONS
        ]
        fig.add_trace(go.Scattergeo(
            lon=[c["lon"] for c in CANDIDATE_STATIONS],
            lat=[c["lat"] for c in CANDIDATE_STATIONS],
            mode="markers",
            marker=dict(size=14, symbol="star", color=CANDIDATE_COLOR, line=dict(width=1, color="#ffffff")),
            hoverinfo="text", hovertext=cand_hover, name="Candidate station",
        ))
        # Dashed connector from each candidate to the node it would replace.
        for c in CANDIDATE_STATIONS:
            if c["replaces"] and c["replaces"] in NODES_RAW.index:
                old = NODES_RAW.loc[c["replaces"]]
                fig.add_trace(go.Scattergeo(
                    lon=[c["lon"], old.longitude], lat=[c["lat"], old.latitude],
                    mode="lines", line=dict(width=1.5, color=CANDIDATE_COLOR, dash="dot"),
                    hoverinfo="skip", showlegend=False,
                ))

    fig.update_layout(uirevision="railway-map")
    return _base_geo_layout(fig)


# ==========================================
# 7B. MAP RENDERING - TRUCK TAB
# ==========================================
TRUCK_BASE_COLOR = "#2980b9"
TRUCK_ADDED_COLOR = "#8e44ad"
TRUCK_FROM_SEL_COLOR = "#27ae60"
TRUCK_TO_SEL_COLOR = "#e74c3c"


def generate_truck_map_figure(from_sel=None, to_sel=None):
    df = load_truck_matrix()
    arcs = get_truck_arcs(df)
    added = dashboard_added_truck_arcs()
    fig = go.Figure()

    arrow_lons, arrow_lats, arrow_angles, arrow_hovers, arrow_colors = [], [], [], [], []
    hit_lons, hit_lats, hit_hovers = [], [], []

    for f, t in arcs:
        if f not in NODES_RAW.index or t not in NODES_RAW.index:
            continue
        node_a, node_b = NODES_RAW.loc[f], NODES_RAW.loc[t]
        is_added = (f, t) in added
        color = TRUCK_ADDED_COLOR if is_added else TRUCK_BASE_COLOR

        hover_txt = (
            f"<b>{node_a.node_name}</b> (#{f}) &rarr; <b>{node_b.node_name}</b> (#{t})<br>"
            f"Distance: {df.loc[f, t]:.2f} km<br>"
            f"{'Added via dashboard' if is_added else 'Original (ArcGIS)'}"
        )

        fig.add_trace(go.Scattergeo(
            lon=[node_a.longitude, node_b.longitude], lat=[node_a.latitude, node_b.latitude],
            mode="lines", line=dict(width=3, color=color), hoverinfo="skip",
        ))
        hit_lons += [node_a.longitude, node_b.longitude, None]
        hit_lats += [node_a.latitude, node_b.latitude, None]
        hit_hovers += [hover_txt, hover_txt, hover_txt]

        tip_lat, tip_lon = _arrow_tip(node_a.latitude, node_a.longitude, node_b.latitude, node_b.longitude)
        bearing = _bearing(node_a.latitude, node_a.longitude, node_b.latitude, node_b.longitude)
        arrow_lons.append(tip_lon)
        arrow_lats.append(tip_lat)
        arrow_angles.append(bearing)
        arrow_hovers.append(hover_txt)
        arrow_colors.append(color)

    if hit_lons:
        fig.add_trace(go.Scattergeo(
            lon=hit_lons, lat=hit_lats, mode="lines",
            line=dict(width=14, color="rgba(0,0,0,0.001)"),
            hoverinfo="text", hovertext=hit_hovers, name="Arc",
        ))
    if arrow_lons:
        fig.add_trace(go.Scattergeo(
            lon=arrow_lons, lat=arrow_lats, mode="markers",
            marker=dict(symbol="arrow", size=12, color=arrow_colors, angle=arrow_angles, line=dict(width=0)),
            hoverinfo="text", hovertext=arrow_hovers, name="Direction",
        ))

    # All nodes, colored by type and click-able (customdata=node_id) to pick From/To.
    node_colors = [_node_color(t) for t in NODES_RAW["node_type"]]
    node_hover = [
        f"{row.node_name} (#{nid})<br>Type: {row.node_type}<br>Click to set as From/To"
        for nid, row in NODES_RAW.iterrows()
    ]
    fig.add_trace(go.Scattergeo(
        lon=NODES_RAW["longitude"], lat=NODES_RAW["latitude"],
        mode="markers+text",
        text=NODES_RAW.index.astype(str),
        textposition="top right",
        marker=dict(size=10, color=node_colors, line=dict(width=1.2, color="#ffffff")),
        hoverinfo="text", hovertext=node_hover,
        customdata=[int(nid) for nid in NODES_RAW.index], name="Nodes",
    ))

    # Ring the current From/To selection so map clicks and dropdowns stay in sync visually.
    for sel, label, color in [(from_sel, "From", TRUCK_FROM_SEL_COLOR), (to_sel, "To", TRUCK_TO_SEL_COLOR)]:
        if sel is not None and sel in NODES_RAW.index:
            row = NODES_RAW.loc[sel]
            fig.add_trace(go.Scattergeo(
                lon=[row.longitude], lat=[row.latitude], mode="markers",
                marker=dict(size=24, color="rgba(0,0,0,0)", line=dict(width=3, color=color)),
                hoverinfo="text", hovertext=[f"{label}: {row.node_name}"], name=label,
            ))

    fig.update_layout(uirevision="truck-map")
    return _base_geo_layout(fig)


# ==========================================
# 8. APP LAYOUT
# ==========================================
app = dash.Dash(__name__, external_stylesheets=[dbc.themes.BOOTSTRAP])
app.title = "Network Connections Dashboard"

CLASS_OPTIONS = [
    {
        "label": f" {sc.capitalize()}  ({SIZE_CLASS_RANGES_KG_S[sc][0]:.1f}-{SIZE_CLASS_RANGES_KG_S[sc][1]:.1f} kg/s, "
                  f"≤{SIZE_CLASS_MAX_CAPACITY_T_H[sc]:,.0f} t/h)",
        "value": sc,
    }
    for sc in SIZE_CLASSES
]

PIPELINE_LEGEND = dbc.Row([
    dbc.Col(html.Span("─ ", style={"color": ENABLED_SOLID, "fontWeight": "bold"}), width="auto"),
    dbc.Col("Enabled for selected class", width="auto", className="me-3"),
    dbc.Col(html.Span("─ ", style={"color": DISABLED_SOLID, "fontWeight": "bold"}), width="auto"),
    dbc.Col("Disabled for selected class", width="auto"),
], className="small text-muted mb-2")

PIPELINE_TAB_CONTENT = html.Div([
    dcc.Store(id="data-version", data=0),
    dbc.Row([dbc.Col(PIPELINE_LEGEND, width=12)]),
    dbc.Row([
        dbc.Col([
            dcc.Graph(id="network-map", style={"height": "70vh"}),
        ], width=7),
        dbc.Col([
            dbc.Card([
                dbc.CardBody([
                    html.H5("Size class (for map click + bulk actions)", className="card-title"),
                    dcc.RadioItems(
                        id="class-selector",
                        options=CLASS_OPTIONS,
                        value="small",
                        labelStyle={"display": "block", "margin-bottom": "6px"},
                        className="mb-2",
                    ),
                    html.Div(id="class-status-text", className="small text-muted mb-3"),
                    dbc.Row([
                        dbc.Col(dbc.Button("Enable all for this class", id="enable-all-btn",
                                            color="success", outline=True, size="sm", className="w-100"), width=6),
                        dbc.Col(dbc.Button("Disable all for this class", id="disable-all-btn",
                                            color="danger", outline=True, size="sm", className="w-100"), width=6),
                    ], className="mb-2"),
                    html.Div(id="bulk-status-text", className="small mt-1"),
                    html.Hr(),
                    html.P(
                        "Tip: click a ✅/❌ cell in the Small/Medium/Large columns below to toggle "
                        "that arc for that specific class directly, regardless of which class is selected "
                        "above. Click an arc on the map to toggle it for the selected class.",
                        className="small text-muted",
                    ),
                ])
            ], style={"height": "70vh", "overflowY": "auto"}),
        ], width=5),
    ]),
    html.Hr(),
    html.H5("All pipeline arcs"),
    dash_table.DataTable(
        id="arcs-table",
        columns=[
            {"name": "From", "id": "from_name"}, {"name": "id", "id": "from_node"},
            {"name": "To", "id": "to_name"}, {"name": "id", "id": "to_node"},
            {"name": "Distance (km)", "id": "distance_km"},
            {"name": "From-node flow (kg/s)", "id": "from_flow_kg_s"},
            {"name": "Small", "id": "small"},
            {"name": "Medium", "id": "medium"},
            {"name": "Large", "id": "large"},
        ],
        data=build_arcs_table_data(),
        style_table={"overflowX": "auto"},
        style_cell={"fontSize": 12, "padding": "4px", "textAlign": "center"},
        style_cell_conditional=[
            {"if": {"column_id": c}, "textAlign": "left"} for c in ["from_name", "to_name"]
        ],
        style_data_conditional=[
            {"if": {"column_id": sc}, "cursor": "pointer"} for sc in SIZE_CLASSES
        ],
        sort_action="native",
        filter_action="native",
        page_size=15,
    ),
], className="pt-3")

RAIL_LEGEND = dbc.Row([
    dbc.Col(html.Span("─ ", style={"color": BIDIRECTIONAL_COLOR, "fontWeight": "bold"}), width="auto"),
    dbc.Col("Bidirectional arc", width="auto", className="me-3"),
    dbc.Col(html.Span("─ ", style={"color": ONEWAY_COLOR, "fontWeight": "bold"}), width="auto"),
    dbc.Col("One-way only (reverse missing)", width="auto", className="me-3"),
    dbc.Col(html.Span("◯ ", style={"color": ORPHAN_RING_COLOR, "fontWeight": "bold"}), width="auto"),
    dbc.Col("Orphaned station (0 arcs)", width="auto", className="me-3"),
    dbc.Col(html.Span("★ ", style={"color": CANDIDATE_COLOR, "fontWeight": "bold"}), width="auto"),
    dbc.Col("Candidate replacement station", width="auto"),
], className="small text-muted mb-2 flex-wrap")

n_transport = int((NODES_RAW["node_type"].str.contains("Transport")).sum())
n_orphaned = len(ORPHANED_TRANSPORT_NODES)
n_bidir = sum(1 for f, t in POSSIBLE_RAIL_ARCS if (t, f) in RAIL_ARC_SET)
n_oneway = len(POSSIBLE_RAIL_ARCS) - n_bidir

RAIL_TAB_CONTENT = html.Div([
    dbc.Row([dbc.Col(RAIL_LEGEND, width=12)]),
    dbc.Row([
        dbc.Col([
            dcc.Checklist(
                id="show-candidates-checkbox",
                options=[{"label": " Show candidate replacement stations (from station audit)", "value": "show"}],
                value=[],
                className="small mb-2",
            ),
            dcc.Graph(id="rail-map", style={"height": "70vh"}),
        ], width=7),
        dbc.Col([
            dbc.Card([
                dbc.CardBody([
                    html.H5("Railway network summary", className="card-title"),
                    html.Ul([
                        html.Li(f"'Transport' (station) nodes defined: {n_transport}"),
                        html.Li(f"Orphaned stations (0 rail arcs): {n_orphaned}", className="text-danger" if n_orphaned else ""),
                        html.Li(f"Directed arcs total: {len(POSSIBLE_RAIL_ARCS)}"),
                        html.Li(f"Bidirectional arcs: {n_bidir}"),
                        html.Li(f"One-way-only arcs (reverse missing): {n_oneway}", className="text-warning" if n_oneway else ""),
                    ], className="small"),
                    html.Hr(),
                    html.H6("Orphaned stations", className="mb-1"),
                    html.Ul(build_orphaned_stations_list(), className="small") if n_orphaned else
                        html.P("None.", className="small text-muted"),
                    html.Hr(),
                    html.P(
                        "This tab is read-only: node_metrics_150.xlsx is hand-maintained and is never "
                        "written to from here. See train_stations_analysis/README.md for the full audit "
                        "behind the candidate stations and orphan/one-way findings.",
                        className="small text-muted",
                    ),
                ])
            ], style={"height": "70vh", "overflowY": "auto"}),
        ], width=5),
    ]),
    html.Hr(),
    html.H5("All railway arcs"),
    dash_table.DataTable(
        id="rail-arcs-table",
        columns=[
            {"name": "From", "id": "from_name"}, {"name": "id", "id": "from_node"},
            {"name": "To", "id": "to_name"}, {"name": "id", "id": "to_node"},
            {"name": "Distance (km)", "id": "distance_km"},
            {"name": "Direction", "id": "direction"},
            {"name": "From type", "id": "from_type"},
            {"name": "To type", "id": "to_type"},
        ],
        data=build_rail_arcs_table_data(),
        style_table={"overflowX": "auto"},
        style_cell={"fontSize": 12, "padding": "4px", "textAlign": "center"},
        style_cell_conditional=[
            {"if": {"column_id": c}, "textAlign": "left"} for c in ["from_name", "to_name"]
        ],
        style_data_conditional=[
            {"if": {"filter_query": '{direction} = "One-way only"'}, "backgroundColor": "#fdf2e9"},
        ],
        sort_action="native",
        filter_action="native",
        page_size=15,
    ),
], className="pt-3")

TRUCK_LEGEND = dbc.Row([
    dbc.Col(html.Span("─ ", style={"color": TRUCK_BASE_COLOR, "fontWeight": "bold"}), width="auto"),
    dbc.Col("Original (ArcGIS) arc", width="auto", className="me-3"),
    dbc.Col(html.Span("─ ", style={"color": TRUCK_ADDED_COLOR, "fontWeight": "bold"}), width="auto"),
    dbc.Col("Added via dashboard", width="auto", className="me-3"),
    dbc.Col(html.Span("◯ ", style={"color": TRUCK_FROM_SEL_COLOR, "fontWeight": "bold"}), width="auto"),
    dbc.Col("Selected 'From'", width="auto", className="me-3"),
    dbc.Col(html.Span("◯ ", style={"color": TRUCK_TO_SEL_COLOR, "fontWeight": "bold"}), width="auto"),
    dbc.Col("Selected 'To'", width="auto"),
], className="small text-muted mb-2 flex-wrap")

TRUCK_TAB_CONTENT = html.Div([
    dcc.Store(id="truck-data-version", data=0),
    dcc.Store(id="truck-pending-save", data=None),
    dbc.Row([dbc.Col(TRUCK_LEGEND, width=12)]),
    dbc.Row([
        dbc.Col([
            dcc.Graph(id="truck-map", style={"height": "65vh"}),
        ], width=7),
        dbc.Col([
            dbc.Card([
                dbc.CardBody([
                    html.H5("Add a new truck arc", className="card-title"),
                    html.P(
                        "Pick two nodes - via the dropdowns or by clicking their markers on the map "
                        "(fills From, then To; use Clear to start over).",
                        className="small text-muted",
                    ),
                    dbc.Row([
                        dbc.Col([
                            html.Label("From", className="small"),
                            dcc.Dropdown(id="truck-from-dropdown", options=NODE_DROPDOWN_OPTIONS, placeholder="From node"),
                        ], width=6),
                        dbc.Col([
                            html.Label("To", className="small"),
                            dcc.Dropdown(id="truck-to-dropdown", options=NODE_DROPDOWN_OPTIONS, placeholder="To node"),
                        ], width=6),
                    ], className="mb-2"),
                    dbc.Button("Clear selection", id="truck-clear-btn", size="sm", color="secondary",
                               outline=True, className="mb-3"),

                    dbc.Button("Compute via OSM routing", id="truck-compute-btn", color="primary",
                               size="sm", className="w-100 mb-2"),
                    dcc.Checklist(
                        id="truck-reverse-checkbox",
                        options=[{"label": " Also compute/save the reverse direction", "value": "rev"}],
                        value=[], className="small mb-2",
                    ),
                    dcc.Loading(html.Div(id="truck-compute-status", className="small mb-2"), type="dot"),

                    dbc.Row([
                        dbc.Col([
                            html.Label("Distance from→to (km)", className="small"),
                            dcc.Input(id="truck-distance-input", type="number", min=0,
                                      className="form-control form-control-sm"),
                        ], width=6),
                        dbc.Col([
                            html.Label("Distance to→from (km)", className="small"),
                            dcc.Input(id="truck-reverse-distance-input", type="number", min=0,
                                      className="form-control form-control-sm"),
                        ], width=6),
                    ], className="mb-2"),

                    dbc.Button("Save arc", id="truck-save-btn", color="success", size="sm", className="w-100 mb-2"),
                    dbc.Button("Confirm overwrite", id="truck-overwrite-btn", color="danger", size="sm",
                               className="w-100 mb-2", style={"display": "none"}),
                    html.Div(id="truck-save-status", className="small"),

                    html.Hr(),
                    html.H6("Summary", className="mb-1"),
                    html.Ul(id="truck-summary", className="small"),
                    html.Hr(),
                    html.P(
                        "Saving writes straight into node_metrics_150.xlsx's 'truck' sheet (only that "
                        "sheet's cell(s) - every other sheet is untouched), so the next main_italy.py run "
                        "picks the new arc up automatically. The file is backed up once per dashboard "
                        "session before the first write.",
                        className="small text-muted",
                    ),
                ])
            ], style={"height": "65vh", "overflowY": "auto"}),
        ], width=5),
    ]),
    html.Hr(),
    html.H5("All truck arcs"),
    dash_table.DataTable(
        id="truck-arcs-table",
        columns=[
            {"name": "From", "id": "from_name"}, {"name": "id", "id": "from_node"},
            {"name": "To", "id": "to_name"}, {"name": "id", "id": "to_node"},
            {"name": "Distance (km)", "id": "distance_km"},
            {"name": "Source", "id": "source"},
            {"name": "Added at", "id": "added_at"},
        ],
        data=build_truck_arcs_table_data(),
        style_table={"overflowX": "auto"},
        style_cell={"fontSize": 12, "padding": "4px", "textAlign": "center"},
        style_cell_conditional=[
            {"if": {"column_id": c}, "textAlign": "left"} for c in ["from_name", "to_name", "source"]
        ],
        style_data_conditional=[
            {"if": {"filter_query": '{source} contains "Dashboard"'}, "backgroundColor": "#f4ecf7"},
        ],
        sort_action="native",
        filter_action="native",
        page_size=15,
    ),
], className="pt-3")

app.layout = dbc.Container([
    dbc.Row([dbc.Col(html.H3("AdOpT-NET0 Italy - Network Connections Dashboard", className="text-center my-3"), width=12)]),
    dbc.Tabs([
        dbc.Tab(PIPELINE_TAB_CONTENT, label="Pipeline size classes", tab_id="pipeline-tab"),
        dbc.Tab(RAIL_TAB_CONTENT, label="Railway network", tab_id="rail-tab"),
        dbc.Tab(TRUCK_TAB_CONTENT, label="Truck network", tab_id="truck-tab"),
    ], id="main-tabs", active_tab="pipeline-tab"),
], fluid=True)


# ==========================================
# 9. CALLBACKS - PIPELINE TAB
# ==========================================
@app.callback(Output("network-map", "figure"),
              Input("class-selector", "value"), Input("data-version", "data"))
def refresh_map(size_class, _version):
    return generate_pipeline_map_figure(size_class)


@app.callback(Output("arcs-table", "data"), Input("data-version", "data"))
def refresh_table(_version):
    return build_arcs_table_data()


@app.callback(Output("class-status-text", "children"), Input("class-selector", "value"), Input("data-version", "data"))
def show_class_status(size_class, _version):
    mask_df = load_class_mask(size_class)
    enabled = sum(1 for f, t in POSSIBLE_ARCS if bool(mask_df.loc[f, t]))
    total = len(POSSIBLE_ARCS)
    return f"'{size_class}': {enabled}/{total} arcs enabled."


@app.callback(
    Output("data-version", "data", allow_duplicate=True),
    Input("network-map", "clickData"),
    State("class-selector", "value"),
    State("data-version", "data"),
    prevent_initial_call=True,
)
def toggle_from_map(click_data, size_class, version):
    if not click_data:
        return dash.no_update
    point = click_data["points"][0]
    if "customdata" not in point or not isinstance(point["customdata"], list):
        return dash.no_update
    from_node, to_node = point["customdata"]
    toggle_arc(size_class, int(from_node), int(to_node))
    return (version or 0) + 1


@app.callback(
    Output("data-version", "data", allow_duplicate=True),
    Input("arcs-table", "active_cell"),
    State("arcs-table", "data"),
    State("data-version", "data"),
    prevent_initial_call=True,
)
def toggle_from_table(active_cell, table_data, version):
    if not active_cell or active_cell["column_id"] not in SIZE_CLASSES:
        return dash.no_update
    row = table_data[active_cell["row"]]
    size_class = active_cell["column_id"]
    toggle_arc(size_class, int(row["from_node"]), int(row["to_node"]))
    return (version or 0) + 1


@app.callback(
    Output("bulk-status-text", "children"),
    Output("data-version", "data", allow_duplicate=True),
    Input("enable-all-btn", "n_clicks"),
    Input("disable-all-btn", "n_clicks"),
    State("class-selector", "value"),
    State("data-version", "data"),
    prevent_initial_call=True,
)
def bulk_toggle(_enable_clicks, _disable_clicks, size_class, version):
    triggered = dash.ctx.triggered_id
    if triggered == "enable-all-btn":
        set_all_arcs(size_class, 1)
        msg = f"✅ All arcs enabled for '{size_class}'."
    elif triggered == "disable-all-btn":
        set_all_arcs(size_class, 0)
        msg = f"\U0001F5D1️ All arcs disabled for '{size_class}'."
    else:
        return dash.no_update, dash.no_update
    return msg, (version or 0) + 1


# ==========================================
# 10. CALLBACKS - RAILWAY TAB
# ==========================================
@app.callback(Output("rail-map", "figure"), Input("show-candidates-checkbox", "value"))
def refresh_rail_map(checkbox_value):
    return generate_rail_map_figure(show_candidates=bool(checkbox_value))


# ==========================================
# 11. CALLBACKS - TRUCK TAB
# ==========================================
@app.callback(
    Output("truck-map", "figure"),
    Input("truck-data-version", "data"),
    Input("truck-from-dropdown", "value"),
    Input("truck-to-dropdown", "value"),
)
def refresh_truck_map(_version, from_sel, to_sel):
    return generate_truck_map_figure(from_sel, to_sel)


@app.callback(Output("truck-arcs-table", "data"), Input("truck-data-version", "data"))
def refresh_truck_table(_version):
    return build_truck_arcs_table_data()


@app.callback(Output("truck-summary", "children"), Input("truck-data-version", "data"))
def refresh_truck_summary(_version):
    arcs = get_truck_arcs()
    added = dashboard_added_truck_arcs()
    n_added = sum(1 for a in arcs if a in added)
    return [
        html.Li(f"Directed arcs total: {len(arcs)}"),
        html.Li(f"Original (ArcGIS) arcs: {len(arcs) - n_added}"),
        html.Li(f"Added via dashboard: {n_added}"),
    ]


@app.callback(
    Output("truck-from-dropdown", "value"),
    Output("truck-to-dropdown", "value"),
    Input("truck-map", "clickData"),
    Input("truck-clear-btn", "n_clicks"),
    State("truck-from-dropdown", "value"),
    State("truck-to-dropdown", "value"),
    prevent_initial_call=True,
)
def pick_or_clear_truck_nodes(click_data, _clear_clicks, from_val, to_val):
    if dash.ctx.triggered_id == "truck-clear-btn":
        return None, None

    if not click_data:
        return dash.no_update, dash.no_update
    point = click_data["points"][0]
    cd = point.get("customdata")
    # Node markers carry a scalar node_id; arc/direction traces carry no
    # customdata (or a list, for other tabs' click-to-toggle arcs) - only
    # react to an actual node click here.
    if cd is None or isinstance(cd, list):
        return dash.no_update, dash.no_update
    node_id = int(cd)
    if from_val is None:
        return node_id, dash.no_update
    if to_val is None:
        return dash.no_update, node_id
    return dash.no_update, dash.no_update


@app.callback(
    Output("truck-distance-input", "value"),
    Output("truck-reverse-distance-input", "value"),
    Output("truck-compute-status", "children"),
    Input("truck-compute-btn", "n_clicks"),
    State("truck-from-dropdown", "value"),
    State("truck-to-dropdown", "value"),
    State("truck-reverse-checkbox", "value"),
    prevent_initial_call=True,
)
def do_compute_truck_distance(_n, from_node, to_node, reverse_checked):
    d_fwd, msg_fwd = compute_truck_distance(from_node, to_node)
    d_rev, msg_rev = None, None
    if reverse_checked:
        d_rev, msg_rev = compute_truck_distance(to_node, from_node)
    full_msg = msg_fwd if not reverse_checked else f"Forward: {msg_fwd} | Reverse: {msg_rev}"
    return d_fwd, d_rev, full_msg


@app.callback(
    Output("truck-save-status", "children"),
    Output("truck-data-version", "data", allow_duplicate=True),
    Output("truck-pending-save", "data"),
    Output("truck-overwrite-btn", "style"),
    Input("truck-save-btn", "n_clicks"),
    State("truck-from-dropdown", "value"),
    State("truck-to-dropdown", "value"),
    State("truck-distance-input", "value"),
    State("truck-reverse-checkbox", "value"),
    State("truck-reverse-distance-input", "value"),
    State("truck-data-version", "data"),
    prevent_initial_call=True,
)
def save_truck_arc(_n, from_node, to_node, distance, reverse_checked, reverse_distance, version):
    hide, show = {"display": "none"}, {"display": "block"}

    if from_node is None or to_node is None or distance is None:
        return "Select From and To, then compute or enter a distance first.", dash.no_update, None, hide
    if from_node == to_node:
        return "From and To must be different nodes.", dash.no_update, None, hide

    want_reverse = bool(reverse_checked) and reverse_distance is not None
    current_fwd = get_current_truck_value(from_node, to_node)
    current_rev = get_current_truck_value(to_node, from_node) if want_reverse else None
    fwd_conflict = current_fwd not in (None, 0)
    rev_conflict = want_reverse and current_rev not in (None, 0)

    if fwd_conflict or rev_conflict:
        msg = "⚠️ Existing value(s) would be overwritten - "
        if fwd_conflict:
            msg += f"{NODE_NAME.get(from_node)}→{NODE_NAME.get(to_node)}: {current_fwd} km → {distance} km. "
        if rev_conflict:
            msg += f"{NODE_NAME.get(to_node)}→{NODE_NAME.get(from_node)}: {current_rev} km → {reverse_distance} km. "
        msg += "Click 'Confirm overwrite' to proceed."
        pending = {"from": from_node, "to": to_node, "distance": distance,
                   "reverse": want_reverse, "reverse_distance": reverse_distance}
        return msg, dash.no_update, pending, show

    write_truck_arc(from_node, to_node, distance, method="dashboard_new")
    if want_reverse:
        write_truck_arc(to_node, from_node, reverse_distance, method="dashboard_new")
    return "✅ Saved to node_metrics_150.xlsx.", (version or 0) + 1, None, hide


@app.callback(
    Output("truck-save-status", "children", allow_duplicate=True),
    Output("truck-data-version", "data", allow_duplicate=True),
    Output("truck-pending-save", "data", allow_duplicate=True),
    Output("truck-overwrite-btn", "style", allow_duplicate=True),
    Input("truck-overwrite-btn", "n_clicks"),
    State("truck-pending-save", "data"),
    State("truck-data-version", "data"),
    prevent_initial_call=True,
)
def confirm_overwrite_truck_arc(_n, pending, version):
    hide = {"display": "none"}
    if not pending:
        return dash.no_update, dash.no_update, dash.no_update, hide
    write_truck_arc(pending["from"], pending["to"], pending["distance"], method="dashboard_overwrite")
    if pending.get("reverse") and pending.get("reverse_distance") is not None:
        write_truck_arc(pending["to"], pending["from"], pending["reverse_distance"], method="dashboard_overwrite")
    return "✅ Overwritten and saved to node_metrics_150.xlsx.", (version or 0) + 1, None, hide


if __name__ == "__main__":
    # threaded=True matters here specifically: "Compute via OSM routing" can
    # block for a long time (tens of seconds to a few minutes for a distant
    # node pair - see truck_routing.py's own per-arc retry/timeout comments).
    # Without it, the single-threaded dev server would freeze every other
    # tab/user for the whole app while one such request is in flight.
    app.run(debug=True, port=8052, threaded=True)
