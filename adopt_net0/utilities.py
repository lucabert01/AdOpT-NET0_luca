from pyomo.environ import SolverFactory
import json
import pandas as pd
from pathlib import Path


def get_gurobi_parameters(solveroptions: dict):
    """
    Initiates the gurobi solver and defines solver parameters

    :param dict solveroptions: dict with solver parameters
    :return: Gurobi Solver
    """
    solver = SolverFactory(solveroptions["solver"]["value"], solver_io="python")
    solver.options["TimeLimit"] = solveroptions["timelim"]["value"] * 3600
    solver.options["MIPGap"] = solveroptions["mipgap"]["value"]
    solver.options["MIPFocus"] = solveroptions["mipfocus"]["value"]
    solver.options["Threads"] = solveroptions["threads"]["value"]
    solver.options["NodefileStart"] = solveroptions["nodefilestart"]["value"]
    solver.options["Method"] = solveroptions["method"]["value"]
    solver.options["Heuristics"] = solveroptions["heuristics"]["value"]
    solver.options["Presolve"] = solveroptions["presolve"]["value"]
    solver.options["BranchDir"] = solveroptions["branchdir"]["value"]
    solver.options["LPWarmStart"] = solveroptions["lpwarmstart"]["value"]
    solver.options["IntFeasTol"] = solveroptions["intfeastol"]["value"]
    solver.options["FeasibilityTol"] = solveroptions["feastol"]["value"]
    solver.options["Cuts"] = solveroptions["cuts"]["value"]
    solver.options["NumericFocus"] = solveroptions["numericfocus"]["value"]

    return solver


def get_glpk_parameters(solveroptions: dict):
    """
    Initiates the glpk solver and defines solver parameters

    :param dict solveroptions: dict with solver parameters
    :return: Gurobi Solver
    """
    solver = SolverFactory("glpk")

    return solver


def get_set_t(config: dict, model_block):
    """
    Returns the correct set_t for different clustering options

    :param dict config: config dict
    :param model_block: pyomo block holding set_t_full and set_t_clustered
    :return: set_t
    """
    if config["optimization"]["typicaldays"]["N"]["value"] == 0:
        return model_block.set_t_full
    elif config["optimization"]["typicaldays"]["method"]["value"] == 1:
        return model_block.set_t_clustered
    elif config["optimization"]["typicaldays"]["method"]["value"] == 2:
        return model_block.set_t_full


def get_data_for_investment_period(
    data, investment_period: str, aggregation_model: str
) -> dict:
    """
    Gets data from DataHandle for specific investement_period. Writes it to a dict.

    :param data: data to use
    :param str investment_period: investment period
    :param str aggregation_model: aggregation type
    :return: data of respective investment period
    :rtype: dict
    """
    data_period = {}
    data_period["period_name"] = investment_period
    data_period["topology"] = data.topology
    data_period["technology_data"] = data.technology_data[investment_period]
    data_period["time_series"] = data.time_series[aggregation_model].loc[
        :, investment_period
    ]
    data_period["network_data"] = data.network_data[investment_period]
    data_period["energybalance_options"] = data.energybalance_options[investment_period]
    data_period["config"] = data.model_config
    if data.model_config["optimization"]["typicaldays"]["N"]["value"] != 0:
        data_period["k_means_specs"] = data.k_means_specs[investment_period]
        # data_period["averaged_specs"] = data.averaged_specs[investment_period]
    if data.model_config["performance"]["pressure"]["pressure_on"]["value"] == 1:
        data_period["compressor_data"] = data.compressor_data[investment_period]

    # Hour multiplication factors
    if data.model_config["optimization"]["typicaldays"]["N"]["value"] == 0:
        data_period["hour_factors"] = [1] * len(
            data_period["topology"]["time_index"]["full"]
        )
    elif data.model_config["optimization"]["typicaldays"]["method"]["value"] == 1:
        data_period["hour_factors"] = data_period["k_means_specs"]["factors"]
    elif data.model_config["optimization"]["typicaldays"]["method"]["value"] == 2:
        data_period["hour_factors"] = [1] * len(
            data_period["topology"]["time_index"]["full"]
        )

    # Nr timesteps averaged
    if data.model_config["optimization"]["timestaging"]["value"] != 0:
        data_period["nr_timesteps_averaged"] = data.model_config["optimization"][
            "timestaging"
        ]["value"]
    else:
        data_period["nr_timesteps_averaged"] = 1

    return data_period


def determine_flow_existing_compressors(self, compressor, b_period, node):
    """
    Determines the flow capacity of an existing compressor connection by returning
    the minimum available capacity between the output and input components

    :param compressor: tuple with carrier, component1, component 2
    :param b_period: pyomo block data for period
    :param node: pyomo block data for node
    :return float: minimum capacity between input and output component
    """
    component_output_bound = float("inf")
    component_input_bound = float("inf")
    period_name = b_period.name.split("[")[-1].rstrip("]")
    type_component = [compressor.output_type, compressor.input_type]

    if type_component[0] == "Technology":
        var_output = (
            b_period.node_blocks[node]
            .tech_blocks_active[compressor.output_component]
            .var_output
        )
        component_output_bound = max(var_output[idx].ub for idx in var_output)
    elif type_component[0] == "Network":
        component_output_bound = next(
            iter(
                b_period.network_block[
                    compressor.output_component
                ].para_size_initial.values()
            )
        )
    elif type_component[0] == "Import":
        component_output_bound = max(
            self.data.time_series["full"][period_name][node]["CarrierData"][
                compressor.carrier
            ]["Import limit"]
        )

    elif type_component[0] == "Generic production":
        component_output_bound = max(
            self.data.time_series["full"][period_name][node]["CarrierData"][
                compressor.carrier
            ]["Generic production"]
        )

    if type_component[1] == "Technology":
        var_output = (
            b_period.node_blocks[node]
            .tech_blocks_active[compressor.input_component]
            .var_output
        )
        component_input_bound = max(var_output[idx].ub for idx in var_output)
    elif type_component[1] == "Network":
        component_input_bound = next(
            iter(
                b_period.network_block[
                    compressor.input_component
                ].para_size_initial.values()
            )
        )
    elif type_component[1] == "Demand":
        component_input_bound = max(
            self.data.time_series["full"][period_name][node]["CarrierData"][
                compressor.carrier
            ]["Demand"]
        )
    elif type_component[1] == "Export":
        component_input_bound = max(
            self.data.time_series["full"][period_name][node]["CarrierData"][
                compressor.carrier
            ]["Export limit"]
        )

    size = min(component_output_bound, component_input_bound)

    return size


def flatten_json_for_excel(data, parent_key=""):
    """
    Flatten JSON data maintaining hierarchical structure for Excel headers
    Special handling for carrier-specific attributes to avoid too many columns
    """
    flattened = {}

    # Define carrier-specific attributes that should be consolidated
    carrier_attributes = [
        "input_carrier",
        "output_carrier",
        "main_input_carrier",
        "main_output_carrier",
        "carrier",
    ]

    for key, value in data.items():
        # Create hierarchical key
        new_key = f"{parent_key}.{key}" if parent_key else key

        if isinstance(value, dict):
            # Check if this is a carrier-specific attribute with units (Units section)
            if any(attr in key.lower() for attr in carrier_attributes) and all(
                isinstance(v, (str, int, float)) for v in value.values()
            ):
                # Consolidate carrier values into a single string: "electricity: MW; hydrogen: kg/h"
                carrier_values = []
                for carrier, unit in value.items():
                    carrier_values.append(f"{carrier}: {unit}")
                flattened[new_key] = "; ".join(carrier_values)
            else:
                # Recursively flatten nested dictionaries
                nested_flattened = flatten_json_for_excel(value, new_key)
                flattened.update(nested_flattened)
        elif isinstance(value, list):
            # Handle carrier lists (Performance section) and regular lists
            if any(attr in key.lower() for attr in carrier_attributes):
                # For carrier lists, join with commas: "electricity, hydrogen"
                flattened[new_key] = ", ".join(str(v) for v in value)
            else:
                # Convert other lists to comma-separated strings
                flattened[new_key] = ", ".join(str(v) for v in value)
        else:
            # Direct value
            flattened[new_key] = value

    return flattened


def create_hierarchical_headers(columns):
    """
    Create hierarchical headers for Excel from flattened column names
    Returns tuple of (header_level_1, header_level_2, has_hierarchy)
    """
    level1_headers = []
    level2_headers = []
    has_hierarchy = False

    for col in columns:
        if "." in col:
            parts = col.split(".", 1)
            level1_headers.append(parts[0])
            level2_headers.append(parts[1])
            has_hierarchy = True
        else:
            level1_headers.append(col)
            level2_headers.append("")

    return level1_headers, level2_headers, has_hierarchy


def create_csv_database_from_json():
    """
    Creates an Excel database from all JSON files in the database/templates directory.
    Creates separate sheets for Technologies and Networks with hierarchical headers.
    """
    # Define paths
    base_path = Path(__file__).parent / "database" / "templates"
    output_path = (
        Path(__file__).parent / "database" / "data" / "Components_database.xlsx"
    )

    # Ensure output directory exists
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Initialize data containers
    tech_data = []
    network_data = []

    # Process technology data files
    tech_data_path = base_path / "technology_data"
    if tech_data_path.exists():
        print(f"Processing technology data from: {tech_data_path}")

        for json_file in tech_data_path.rglob("*.json"):
            try:
                print(f"Processing technology: {json_file}")
                with open(json_file, "r", encoding="utf-8") as f:
                    data = json.load(f)

                # Flatten the JSON data
                flattened_data = flatten_json_for_excel(data)

                # Add metadata
                flattened_data["File_Name"] = json_file.name
                flattened_data["Technology_Group"] = json_file.parent.name

                tech_data.append(flattened_data)

            except Exception as e:
                print(f"Error processing {json_file}: {str(e)}")
                continue

    # Process network data files
    network_data_path = base_path / "network_data"
    if network_data_path.exists():
        print(f"Processing network data from: {network_data_path}")

        for json_file in network_data_path.rglob("*.json"):
            try:
                print(f"Processing network: {json_file}")
                with open(json_file, "r", encoding="utf-8") as f:
                    data = json.load(f)

                # Flatten the JSON data
                flattened_data = flatten_json_for_excel(data)

                # Add metadata
                flattened_data["File_Name"] = json_file.name
                flattened_data["Network_Group"] = "Network"

                network_data.append(flattened_data)

            except Exception as e:
                print(f"Error processing {json_file}: {str(e)}")
                continue

    # Create Excel file with hierarchical headers
    if tech_data or network_data:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

            # Create Technologies sheet
            if tech_data:
                tech_df = pd.DataFrame(tech_data)

                # Reorder columns to put metadata first
                meta_cols = ["File_Name", "Technology_Group"]
                other_cols = [col for col in tech_df.columns if col not in meta_cols]
                tech_df = tech_df[meta_cols + other_cols]

                # Create hierarchical headers
                level1_headers, level2_headers, has_hierarchy = (
                    create_hierarchical_headers(tech_df.columns)
                )

                if has_hierarchy:
                    # Write to Excel with hierarchical headers (start at row 3, no pandas headers)
                    tech_df.to_excel(
                        writer,
                        sheet_name="Technologies",
                        index=False,
                        startrow=2,
                        header=False,
                    )

                    # Get worksheet and write hierarchical headers
                    worksheet = writer.sheets["Technologies"]

                    # Create font styles
                    from openpyxl.styles import Font, Alignment

                    bold_font = Font(bold=True)
                    center_alignment = Alignment(horizontal="center", vertical="center")

                    # Write level 1 headers with bold and center formatting
                    for i, header in enumerate(level1_headers):
                        cell = worksheet.cell(row=1, column=i + 1, value=header)
                        cell.font = bold_font
                        cell.alignment = center_alignment

                    # Write level 2 headers with center formatting
                    for i, header in enumerate(level2_headers):
                        if header:  # Only write if there's a second level
                            cell = worksheet.cell(row=2, column=i + 1, value=header)
                            cell.alignment = center_alignment

                    # Merge cells for level 1 headers that span multiple columns
                    current_header = None
                    start_col = 1
                    for i, header in enumerate(level1_headers):
                        if header != current_header:
                            if current_header and i > start_col:
                                # Merge previous header if it spans multiple columns
                                if i - start_col > 1:
                                    worksheet.merge_cells(
                                        start_row=1,
                                        start_column=start_col,
                                        end_row=1,
                                        end_column=i,
                                    )
                            current_header = header
                            start_col = i + 1

                    # Handle the last header group
                    if len(level1_headers) > start_col:
                        if len(level1_headers) - start_col > 0:
                            worksheet.merge_cells(
                                start_row=1,
                                start_column=start_col,
                                end_row=1,
                                end_column=len(level1_headers),
                            )
                else:
                    # Write to Excel with single header row (normal pandas output)
                    tech_df.to_excel(writer, sheet_name="Technologies", index=False)

                    # Format the single header row
                    worksheet = writer.sheets["Technologies"]
                    from openpyxl.styles import Font, Alignment

                    bold_font = Font(bold=True)
                    center_alignment = Alignment(horizontal="center", vertical="center")

                    # Apply formatting to header row
                    for i, header in enumerate(tech_df.columns):
                        cell = worksheet.cell(row=1, column=i + 1)
                        cell.font = bold_font
                        cell.alignment = center_alignment

                print(f"Created Technologies sheet with {len(tech_data)} technologies")

            # Create Networks sheet
            if network_data:
                network_df = pd.DataFrame(network_data)

                # Reorder columns to put metadata first
                meta_cols = ["File_Name", "Network_Group"]
                other_cols = [col for col in network_df.columns if col not in meta_cols]
                network_df = network_df[meta_cols + other_cols]

                # Create hierarchical headers
                level1_headers, level2_headers, has_hierarchy = (
                    create_hierarchical_headers(network_df.columns)
                )

                if has_hierarchy:
                    # Write to Excel with hierarchical headers (start at row 3, no pandas headers)
                    network_df.to_excel(
                        writer,
                        sheet_name="Networks",
                        index=False,
                        startrow=2,
                        header=False,
                    )

                    # Get worksheet and write hierarchical headers
                    worksheet = writer.sheets["Networks"]

                    # Create font styles
                    from openpyxl.styles import Font, Alignment

                    bold_font = Font(bold=True)
                    center_alignment = Alignment(horizontal="center", vertical="center")

                    # Write level 1 headers with bold and center formatting
                    for i, header in enumerate(level1_headers):
                        cell = worksheet.cell(row=1, column=i + 1, value=header)
                        cell.font = bold_font
                        cell.alignment = center_alignment

                    # Write level 2 headers with center formatting
                    for i, header in enumerate(level2_headers):
                        if header:  # Only write if there's a second level
                            cell = worksheet.cell(row=2, column=i + 1, value=header)
                            cell.alignment = center_alignment

                    # Merge cells for level 1 headers that span multiple columns
                    current_header = None
                    start_col = 1
                    for i, header in enumerate(level1_headers):
                        if header != current_header:
                            if current_header and i > start_col:
                                # Merge previous header if it spans multiple columns
                                if i - start_col > 1:
                                    worksheet.merge_cells(
                                        start_row=1,
                                        start_column=start_col,
                                        end_row=1,
                                        end_column=i,
                                    )
                            current_header = header
                            start_col = i + 1

                    # Handle the last header group
                    if len(level1_headers) > start_col:
                        if len(level1_headers) - start_col > 0:
                            worksheet.merge_cells(
                                start_row=1,
                                start_column=start_col,
                                end_row=1,
                                end_column=len(level1_headers),
                            )
                else:
                    # Write to Excel with single header row (normal pandas output)
                    network_df.to_excel(writer, sheet_name="Networks", index=False)

                    # Format the single header row
                    worksheet = writer.sheets["Networks"]
                    from openpyxl.styles import Font, Alignment

                    bold_font = Font(bold=True)
                    center_alignment = Alignment(horizontal="center", vertical="center")

                    # Apply formatting to header row
                    for i, header in enumerate(network_df.columns):
                        cell = worksheet.cell(row=1, column=i + 1)
                        cell.font = bold_font
                        cell.alignment = center_alignment

                print(f"Created Networks sheet with {len(network_data)} networks")

        print(f"Excel database created successfully at: {output_path}")
        return output_path

    else:
        print("No data found to create Excel database!")
        return None
