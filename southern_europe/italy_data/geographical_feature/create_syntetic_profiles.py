"""
Generate synthetic hourly emission profiles (8760 hours) for all emitters
that don't already have real-world data in emission_profile_emitters.xlsx.

Sectors handled:
  - Cement   → flat when running; annual capacity factor sampled ~U(65%, 75%),
               downtime split into 3 maintenance stops (2-6 weeks each) placed in
               late July-August, Christmas/New Year, and spring
  - Waste    → running baseline ±5% noise, tuned so annual capacity factor ~U(85%, 90%);
               fixed 3-week full shutdown + a half-capacity stop of 1-5 weeks (not winter)
  - Refining → flat, single ~5.2-week stop to bring annual capacity factor to ~90%
  - Other    → same as Refining
  - Transport / Storage → skipped

All profiles are scaled so that sum(hourly_profile) == annual_flux (tonnes/year).
annual_flux is read from the "annual_flux" column in node_metrics.xlsx (values in kg).

Output: sheet "synthetic_data" written to emission_profile_emitters.xlsx
Column naming convention: "[Sector - node_name]"
"""

from pathlib import Path
import numpy as np
import pandas as pd
from openpyxl import load_workbook
import matplotlib.pyplot as plt

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE          = Path(r"C:\Users\0954659\PycharmProjects\AdOpT-NET0_luca\southern_europe\italy_data\geographical_feature")
NODES_FILE    = BASE / "node_metrics.xlsx"
EMITTERS_FILE = BASE / "emission_profile_emitters.xlsx"

HOURS = 8760
RNG   = np.random.default_rng(42)          # fixed seed → reproducible results

SKIP_SECTORS = {"Transport", "Storage"}

# ---------------------------------------------------------------------------
# Calendar windows (hour indices, non-leap year)
# ---------------------------------------------------------------------------
SPRING_START,   SPRING_END    =  59 * 24, 151 * 24   # 1 Mar – 31 May
SUMMER_START,   SUMMER_END    = 196 * 24, 243 * 24   # 15 Jul – 31 Aug
XMAS_START                    = 349 * 24              # 15 Dec
XMAS_END                      = (365 + 31) * 24        # 31 Jan (wraps into next year)
NO_WINTER_START, NO_WINTER_END =  59 * 24, 334 * 24  # 1 Mar – 30 Nov  (allowed for Waste stops)

WEEK  = 7  * 24   # 168 h

CEMENT_CF_LOW, CEMENT_CF_HIGH = 0.65, 0.75   # annual capacity factor range
CEMENT_MIN_STOP = 2 * WEEK   # 336 h
CEMENT_MAX_STOP = 6 * WEEK   # 1008 h

WASTE_CF_LOW, WASTE_CF_HIGH = 0.85, 0.90   # annual capacity factor range (vs. nameplate = 1.0)
WASTE_FULL_STOP     = 3 * WEEK   # 504 h, fixed-length full shutdown
WASTE_HALF_STOP_MIN = 1 * WEEK   # 168 h
WASTE_HALF_STOP_MAX = 5 * WEEK   # 840 h

REFINING_CF   = 0.90   # annual capacity factor (vs. nameplate = 1.0)
REFINING_STOP = int(round((1.0 - REFINING_CF) * HOURS))   # ~876 h ~= 5.2 weeks


def _rand_start(window_start: int, window_end: int, stop_len: int) -> int:
    """Uniform random start hour for a stop of `stop_len` hours inside the window."""
    latest = window_end - stop_len
    if latest <= window_start:
        return window_start
    return int(RNG.integers(window_start, latest))


def _zero_range_wrapped(profile: np.ndarray, start: int, length: int) -> None:
    """Zero out `length` hours starting at `start`, wrapping around the year boundary."""
    idx = np.arange(start, start + length) % HOURS
    profile[idx] = 0.0


def _split_stop_lengths_hours(total_hours: float, n: int = 3,
                               min_len: int = CEMENT_MIN_STOP, max_len: int = CEMENT_MAX_STOP,
                               max_tries: int = 2000) -> np.ndarray:
    """Split `total_hours` of downtime into `n` stop lengths, each within [min_len, max_len]."""
    span = max_len - min_len
    budget = min(max(total_hours - n * min_len, 0.0), n * span)
    for _ in range(max_tries):
        weights = RNG.dirichlet(np.full(n, 4.0))   # moderately concentrated -> near-equal splits
        extra = weights * budget
        if np.all(extra <= span):
            return min_len + extra
    return np.full(n, np.clip(total_hours / n, min_len, max_len))


# ---------------------------------------------------------------------------
# Normalized profile generators  (output in [0, 1])
# Scaling to physical units (tonnes/h) is done separately via annual_flux.
# ---------------------------------------------------------------------------

def _norm_cement() -> np.ndarray:
    """
    Flat 1.0 when running.
    Annual capacity factor sampled ~ U(65%, 75%); the resulting downtime is split
    into 3 maintenance stops (each 2-6 weeks), one per window:
      - late July – August
      - Christmas / New Year (wraps around the year boundary)
      - spring (Mar – May)
    """
    profile = np.ones(HOURS)

    capacity_factor = RNG.uniform(CEMENT_CF_LOW, CEMENT_CF_HIGH)
    total_downtime_hours = (1.0 - capacity_factor) * HOURS
    stop_lengths = _split_stop_lengths_hours(total_downtime_hours)
    RNG.shuffle(stop_lengths)   # random assignment of lengths to windows

    windows = [
        (SUMMER_START, SUMMER_END),
        (XMAS_START, XMAS_END),
        (SPRING_START, SPRING_END),
    ]
    for (w_start, w_end), length in zip(windows, stop_lengths):
        length = int(round(length))
        s = _rand_start(w_start, w_end, length)
        _zero_range_wrapped(profile, s, length)

    return profile


def _norm_waste() -> np.ndarray:
    """
    Running hours sit at a baseline level B (± 5% noise), where B is tuned per
    node so the annual capacity factor (relative to nameplate = 1.0) lands in
    U(85%, 90%). Two non-overlapping, non-winter (Mar–Nov) stops:
      - a fixed 3-week full shutdown (0.0)
      - a half-capacity stop (0.5 of nameplate, no noise), length U(1, 5) weeks
    """
    half_len = int(round(RNG.uniform(WASTE_HALF_STOP_MIN, WASTE_HALF_STOP_MAX)))
    full_len = WASTE_FULL_STOP

    placed: list[tuple[int, int]] = []
    for length in (full_len, half_len):
        for _ in range(10_000):                  # retry until non-overlapping
            s = _rand_start(NO_WINTER_START, NO_WINTER_END, length)
            if all(s + length <= s0 or s >= s0 + l0 for s0, l0 in placed):
                placed.append((s, length))
                break
        else:
            placed.append((s, length))            # fallback: rare, window nearly full
    (full_start, _), (half_start, _) = placed

    target_cf = RNG.uniform(WASTE_CF_LOW, WASTE_CF_HIGH)
    running_hours = HOURS - full_len - half_len
    baseline = (target_cf * HOURS - half_len * 0.5) / running_hours

    profile = baseline * RNG.uniform(0.95, 1.05, HOURS)
    profile[full_start:full_start + full_len] = 0.0
    profile[half_start:half_start + half_len] = 0.5

    return profile


def _norm_refining() -> np.ndarray:
    """
    Flat 1.0 when running, one random stop long enough (~5.2 weeks) to bring
    the annual capacity factor (relative to nameplate = 1.0) down to ~90%.
    Used for Refining and Other.
    """
    profile = np.ones(HOURS)
    s = _rand_start(0, HOURS, REFINING_STOP)
    profile[s:s + REFINING_STOP] = 0.0
    return profile


NORM_GENERATORS = {
    "Cement":   _norm_cement,
    "Waste":    _norm_waste,
    "Refining": _norm_refining,
    "Other":    _norm_refining,
}

# ---------------------------------------------------------------------------
# Scaling: normalized profile → tonnes/hour so that annual sum = annual_flux
# ---------------------------------------------------------------------------

def scale_profile(normalized: np.ndarray, annual_flux: float) -> np.ndarray:
    """
    Scale normalized profile (dimensionless) to tonnes/hour.
    sum(result) == annual_flux  [tonnes/year]
    """
    annual_flux_t = annual_flux
    total = normalized.sum()
    if total == 0:
        return np.zeros(HOURS)
    return normalized * (annual_flux_t / total)


# ---------------------------------------------------------------------------
# Plotting Function with Subplots Per Emitter
# ---------------------------------------------------------------------------
def plot_profiles(real_df: pd.DataFrame, synth_df: pd.DataFrame) -> None:
    """Generates three figures (Cement, Waste, Refining/Other), with a separate subplot for each emitter."""

    groups = {
        "Cement": ["Cement"],
        "Waste": ["Waste"],
        "Refining & Others": ["Refining", "Other"]
    }

    hours_axis = np.arange(HOURS)

    for plot_title, sectors in groups.items():
        # Identify all columns belonging to this group
        real_cols = [c for c in real_df.columns if any(c.startswith(f"{s} -") for s in sectors)]
        synth_cols = [c for c in synth_df.columns if any(c.startswith(f"{s} -") for s in sectors)]

        all_emitters = [(c, "REAL") for c in real_cols] + [(c, "SYNTHETIC") for c in synth_cols]
        num_emitters = len(all_emitters)

        if num_emitters == 0:
            continue

        # Dynamically compute grid rows and columns (max 3 columns wide)
        cols_grid = min(3, num_emitters)
        rows_grid = int(np.ceil(num_emitters / cols_grid))

        fig, axes = plt.subplots(rows_grid, cols_grid, figsize=(5 * cols_grid, 3.5 * rows_grid), squeeze=False)
        fig.suptitle(f"Hourly Emission Profiles: {plot_title}", fontsize=16, fontweight='bold', y=0.98)

        for idx, (col_name, data_source) in enumerate(all_emitters):
            r, c = divmod(idx, cols_grid)
            ax = axes[r, c]

            if data_source == "REAL":
                ax.scatter(hours_axis, real_df[col_name], s=0.5, alpha=0.6, color="crimson")
                ax.set_title(f"{col_name}\n[REAL DATA]", color="crimson", fontsize=10, fontweight="bold")
                # Light highlight for real data backgrounds
                ax.set_facecolor('#fff5f5')
            else:
                ax.scatter(hours_axis, synth_df[col_name], s=0.3, alpha=0.5, color="#1f77b4")
                ax.set_title(f"{col_name}\n[SYNTHETIC]", color="#1f77b4", fontsize=10)

            ax.grid(True, linestyle="--", alpha=0.4)
            ax.tick_params(labelsize=8)

            # Label only the edge plots to avoid cluttering
            if r == rows_grid - 1:
                ax.set_xlabel("Hour", fontsize=9)
            if c == 0:
                ax.set_ylabel("tonnes/h", fontsize=9)

        # Hide any unused subplot tiles in the grid matrix
        for idx in range(num_emitters, rows_grid * cols_grid):
            r, c = divmod(idx, cols_grid)
            fig.delaxes(axes[r, c])

        plt.tight_layout()
        plt.show()

    # ---------------------------------------------------------------------------
    # Main
    # ---------------------------------------------------------------------------


def main() -> None:
    # 1. Load node list
    nodes_df = pd.read_excel(NODES_FILE, sheet_name="nodes")
    nodes_df = nodes_df[["node_id", "node_name", "node_type", "annual_flux"]].dropna(
        subset=["node_name", "node_type"]
    )
    nodes_df["node_name"] = nodes_df["node_name"].str.strip()
    nodes_df["node_type"] = nodes_df["node_type"].str.strip()
    nodes_df["annual_flux"] = pd.to_numeric(nodes_df["annual_flux"], errors="coerce").fillna(0.0)

    # 2. Load existing real-world profiles
    wb = load_workbook(EMITTERS_FILE)
    real_sheet = "raw_data"
    real_df = pd.read_excel(
        EMITTERS_FILE,
        sheet_name=real_sheet if real_sheet in wb.sheetnames else 0,
    )
    existing_cols = set(real_df.columns)

    # 3. Build synthetic profiles
    synthetic_data: dict[str, np.ndarray] = {}

    for _, row in nodes_df.iterrows():
        node_name = row["node_name"]
        sector = row["node_type"]
        annual_flux = float(row["annual_flux"])

        if sector in SKIP_SECTORS:
            continue

        col_header = f"{sector} - {node_name}"

        if col_header in existing_cols:
            continue

        if sector not in NORM_GENERATORS:
            print(f"[WARN] Unknown node_type '{sector}' for '{node_name}' — using Other generator")
            sector = "Other"
            col_header = f"{sector} - {node_name}"

        if col_header not in synthetic_data:
            normalized = NORM_GENERATORS[sector]()
            synthetic_data[col_header] = scale_profile(normalized, annual_flux)

    synth_df = pd.DataFrame(synthetic_data)

    # 4. Write synthetic_data sheet
    if not synthetic_data:
        print("No synthetic profiles needed — all nodes already have real-world data.")
    else:
        print(f"Writing {len(synth_df.columns)} synthetic profiles ({HOURS} hours each) …")
        with pd.ExcelWriter(EMITTERS_FILE, engine="openpyxl", mode="a",
                            if_sheet_exists="replace") as writer:
            synth_df.to_excel(writer, sheet_name="synthetic_data", index=False)
        print(f"Done. 'synthetic_data' sheet written to:\n  {EMITTERS_FILE}")

    # 5. Generate and display subplot matrices
    print("\nGenerating subplot profiles...")
    plot_profiles(real_df, synth_df)

    # 6. Coverage summary
    print("\n=== Coverage summary ===")
    records = []
    for _, row in nodes_df.iterrows():
        node_name = row["node_name"]
        sector = row["node_type"]
        if sector in SKIP_SECTORS:
            continue
        col_header = f"{sector} - {node_name}"
        source = "real_data" if col_header in existing_cols else "synthetic_data"
        records.append((node_name, sector, round(row["annual_flux"], 1), source))

    coverage = pd.DataFrame(records, columns=["node_name", "sector", "annual_flux_t", "source"])
    print(coverage.to_string(index=False))
    print(f"\nAll {len(coverage)} relevant nodes covered ✓")


if __name__ == "__main__":
    main()