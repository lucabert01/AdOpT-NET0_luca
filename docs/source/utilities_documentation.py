"""
Utilities used by the documentation build only.
This module contains the code that was previously living in
`adopt_net0.utilities` and used to auto-generate the Components database
Excel file for the docs.

This file lives under `docs/source/` so it's not imported by the package
under normal runtime/tests, keeping PR test coverage unaffected.
"""

import json
from pathlib import Path
import pandas as pd


def flatten_json_for_excel(data, parent_key=""):
    """
    Flatten JSON data maintaining hierarchical structure for Excel headers
    Special handling for carrier-specific attributes to avoid too many columns
    """
    flattened = {}

    # Define carrier-specific attributes that should be consolidated
    carrier_attributes = [
        "input_carrier",
        "output_carrier",
        "main_input_carrier",
        "main_output_carrier",
        "carrier",
    ]

    for key, value in data.items():
        # Create hierarchical key
        new_key = f"{parent_key}.{key}" if parent_key else key

        if isinstance(value, dict):
            # Check if this is a carrier-specific attribute with units (Units section)
            if any(attr in key.lower() for attr in carrier_attributes) and all(
                isinstance(v, (str, int, float)) for v in value.values()
            ):
                # Consolidate carrier values into a single string
                carrier_values = []
                for carrier, unit in value.items():
                    carrier_values.append(f"{carrier}: {unit}")
                flattened[new_key] = "; ".join(carrier_values)
            else:
                # Recursively flatten nested dictionaries
                nested_flattened = flatten_json_for_excel(value, new_key)
                flattened.update(nested_flattened)
        elif isinstance(value, list):
            # Handle carrier lists (Performance section) and regular lists
            if any(attr in key.lower() for attr in carrier_attributes):
                # For carrier lists, join with commas: "electricity, hydrogen"
                flattened[new_key] = ", ".join(str(v) for v in value)
            else:
                # Convert other lists to comma-separated strings
                flattened[new_key] = ", ".join(str(v) for v in value)
        else:
            # Direct value
            flattened[new_key] = value

    return flattened


def create_hierarchical_headers(columns):
    """
    Create hierarchical headers for Excel from flattened column names
    Returns tuple of (header_level_1, header_level_2, has_hierarchy)
    """
    level1_headers = []
    level2_headers = []
    has_hierarchy = False

    for col in columns:
        if "." in col:
            parts = col.split(".", 1)
            level1_headers.append(parts[0])
            level2_headers.append(parts[1])
            has_hierarchy = True
        else:
            level1_headers.append(col)
            level2_headers.append("")

    return level1_headers, level2_headers, has_hierarchy


def create_csv_database_from_json():
    """
    Creates an Excel database from all JSON files in the adopt_net0/database/templates directory.
    Creates separate sheets for Technologies and Networks with hierarchical headers.

    This function is intended to be executed as part of the Sphinx documentation
    build (conf.py). It uses repository-relative paths and therefore lives in
    the docs folder.

    Returns the path to the written Excel file (Path) or None.
    """
    # repo root (three levels up from docs/source)
    repo_root = Path(__file__).parent.parent.parent

    base_path = repo_root / "adopt_net0" / "database" / "templates"
    output_path = (
        repo_root / "adopt_net0" / "database" / "data" / "Components_database.xlsx"
    )

    # Ensure output directory exists
    output_path.parent.mkdir(parents=True, exist_ok=True)

    tech_data = []
    network_data = []

    # Process technology data files
    tech_data_path = base_path / "technology_data"
    if tech_data_path.exists():
        for json_file in tech_data_path.rglob("*.json"):
            try:
                with open(json_file, "r", encoding="utf-8") as f:
                    data = json.load(f)

                flattened_data = flatten_json_for_excel(data)
                flattened_data["File_Name"] = json_file.name
                flattened_data["Technology_Group"] = json_file.parent.name
                tech_data.append(flattened_data)
            except Exception:
                continue

    # Process network data files
    network_data_path = base_path / "network_data"
    if network_data_path.exists():
        for json_file in network_data_path.rglob("*.json"):
            try:
                with open(json_file, "r", encoding="utf-8") as f:
                    data = json.load(f)

                flattened_data = flatten_json_for_excel(data)
                flattened_data["File_Name"] = json_file.name
                flattened_data["Network_Group"] = "Network"
                network_data.append(flattened_data)
            except Exception:
                continue

    if tech_data or network_data:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            # Technologies
            if tech_data:
                tech_df = pd.DataFrame(tech_data)
                meta_cols = ["File_Name", "Technology_Group"]
                other_cols = [col for col in tech_df.columns if col not in meta_cols]
                tech_df = tech_df[meta_cols + other_cols]
                level1_headers, level2_headers, has_hierarchy = (
                    create_hierarchical_headers(tech_df.columns)
                )

                if has_hierarchy:
                    tech_df.to_excel(
                        writer,
                        sheet_name="Technologies",
                        index=False,
                        startrow=2,
                        header=False,
                    )
                    worksheet = writer.sheets["Technologies"]
                    from openpyxl.styles import Font, Alignment

                    bold_font = Font(bold=True)
                    center_alignment = Alignment(horizontal="center", vertical="center")
                    for i, header in enumerate(level1_headers):
                        cell = worksheet.cell(row=1, column=i + 1, value=header)
                        cell.font = bold_font
                        cell.alignment = center_alignment
                    for i, header in enumerate(level2_headers):
                        if header:
                            cell = worksheet.cell(row=2, column=i + 1, value=header)
                            cell.alignment = center_alignment
                    # Merge groups
                    current_header = None
                    start_col = 1
                    for i, header in enumerate(level1_headers):
                        if header != current_header:
                            if current_header and i > start_col:
                                if i - start_col > 1:
                                    worksheet.merge_cells(
                                        start_row=1,
                                        start_column=start_col,
                                        end_row=1,
                                        end_column=i,
                                    )
                            current_header = header
                            start_col = i + 1
                    if len(level1_headers) > start_col:
                        if len(level1_headers) - start_col > 0:
                            worksheet.merge_cells(
                                start_row=1,
                                start_column=start_col,
                                end_row=1,
                                end_column=len(level1_headers),
                            )
                else:
                    tech_df.to_excel(writer, sheet_name="Technologies", index=False)
                    worksheet = writer.sheets["Technologies"]
                    from openpyxl.styles import Font, Alignment

                    bold_font = Font(bold=True)
                    center_alignment = Alignment(horizontal="center", vertical="center")
                    for i, header in enumerate(tech_df.columns):
                        cell = worksheet.cell(row=1, column=i + 1)
                        cell.font = bold_font
                        cell.alignment = center_alignment

            # Networks
            if network_data:
                network_df = pd.DataFrame(network_data)
                meta_cols = ["File_Name", "Network_Group"]
                other_cols = [col for col in network_df.columns if col not in meta_cols]
                network_df = network_df[meta_cols + other_cols]
                level1_headers, level2_headers, has_hierarchy = (
                    create_hierarchical_headers(network_df.columns)
                )

                if has_hierarchy:
                    network_df.to_excel(
                        writer,
                        sheet_name="Networks",
                        index=False,
                        startrow=2,
                        header=False,
                    )
                    worksheet = writer.sheets["Networks"]
                    from openpyxl.styles import Font, Alignment

                    bold_font = Font(bold=True)
                    center_alignment = Alignment(horizontal="center", vertical="center")
                    for i, header in enumerate(level1_headers):
                        cell = worksheet.cell(row=1, column=i + 1, value=header)
                        cell.font = bold_font
                        cell.alignment = center_alignment
                    for i, header in enumerate(level2_headers):
                        if header:
                            cell = worksheet.cell(row=2, column=i + 1, value=header)
                            cell.alignment = center_alignment
                    current_header = None
                    start_col = 1
                    for i, header in enumerate(level1_headers):
                        if header != current_header:
                            if current_header and i > start_col:
                                if i - start_col > 1:
                                    worksheet.merge_cells(
                                        start_row=1,
                                        start_column=start_col,
                                        end_row=1,
                                        end_column=i,
                                    )
                            current_header = header
                            start_col = i + 1
                    if len(level1_headers) > start_col:
                        if len(level1_headers) - start_col > 0:
                            worksheet.merge_cells(
                                start_row=1,
                                start_column=start_col,
                                end_row=1,
                                end_column=len(level1_headers),
                            )
                else:
                    network_df.to_excel(writer, sheet_name="Networks", index=False)
                    worksheet = writer.sheets["Networks"]
                    from openpyxl.styles import Font, Alignment

                    bold_font = Font(bold=True)
                    center_alignment = Alignment(horizontal="center", vertical="center")
                    for i, header in enumerate(network_df.columns):
                        cell = worksheet.cell(row=1, column=i + 1)
                        cell.font = bold_font
                        cell.alignment = center_alignment

        return output_path

    return None
